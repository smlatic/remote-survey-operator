import tkinter as tk
from tkinter import ttk
from tkinter import filedialog, messagebox
from openpyxl import load_workbook, Workbook
from collections import defaultdict
import os

def select_file():
    filename = filedialog.askopenfilename(defaultextension=".txt", filetypes=[("Text files", "*.txt")])
    file_entry.delete(0, tk.END)
    file_entry.insert(0, filename)

def main_program():
    filename = file_entry.get()
    append_data = append_var.get()

    if not filename:
        messagebox.showerror("Error", "Please select an input file")
        return

    if not os.path.exists(filename):
        messagebox.showerror("Error", "The file doesn't exist")
        return

    if append_data:
        output_filename = filedialog.askopenfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
        if not output_filename:
            return
        append_to_existing = True
    else:
        output_filename = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
        if not output_filename:
            return
        append_to_existing = False

    with open(filename, 'r') as f:
        lines = f.readlines()

    lines = [line.strip() for line in lines if line.strip() != '']

    if append_to_existing:
        wb = load_workbook(output_filename)  # load existing workbook
    else:
        wb = Workbook()
        wb.remove(wb.active)  # remove default sheet

    data_dict = defaultdict(lambda: defaultdict(list))

    for line in lines:
        if ',' in line:  # this line contains data
            name, layer = line.split(',')
            borehole_num = '_'.join(name.split('_')[:2])  # derive borehole number
            data_dict[layer][borehole_num].append(name)

    summary_data = []  # to store summary data

    for sheet_name, data in data_dict.items():
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.create_sheet(sheet_name)

        for _ in range(5):
            ws.append([])

        max_length = max(len(col) for col in data.values())

        total_lines = sum(len(col) for col in data.values())

        ws.append([f"{sheet_name} Borehole", "total lines", str(total_lines), "Completed", "0", "Remaining", str(total_lines)])

        for i in range(max_length+1):
            row = []
            for borehole_num in sorted(data.keys()):
                if i == 0:
                    row.append(borehole_num)
                    row.append(None)
                else:
                    if i <= len(data[borehole_num]):
                        row.append(data[borehole_num][i-1])
                        row.append(None)
                    else:
                        row.append("")
                        row.append(None)
            ws.append(row)

        for borehole_num in sorted(data.keys()):
            ws.append([borehole_num + " Num lines", len(data[borehole_num]), "0", "", "", "", ""])

        summary_data.append((sheet_name, len(data.keys()), total_lines))  # add summary data

    if 'Summary' in wb.sheetnames:
        ws_summary = wb['Summary']
    else:
        ws_summary = wb.create_sheet('Summary', 0)

    ws_summary.append(["File: " + os.path.basename(filename), "", ""])  # add header for the file

    # Add headers for the summary data
    ws_summary.append(["Sheet Name", "Points of interest", "Total Lines"])

    # write the new summary data
    for row in summary_data:
        ws_summary.append(list(row))

    # Add 3 empty rows after each summary
    for _ in range(3):
        ws_summary.append([])

    wb.save(output_filename)
    messagebox.showinfo("Success", f"File saved successfully as {output_filename}")

root = tk.Tk()
root.title("Excel File Creator")

file_label = tk.Label(root, text="Select the input text file: ")
file_entry = tk.Entry(root, width=50)
file_button = tk.Button(root, text="Browse", command=select_file)
append_label = tk.Label(root, text="Do you want to append to an existing file?")
append_var = tk.BooleanVar()
append_check = tk.Checkbutton(root, variable=append_var)
run_button = tk.Button(root, text="Run", command=main_program)

file_label.grid(row=0, column=0, sticky="e")
file_entry.grid(row=0, column=1)
file_button.grid(row=0, column=2)
append_label.grid(row=1, column=0, sticky="e")
append_check.grid(row=1, column=1, sticky="w")
run_button.grid(row=2, column=0, columnspan=3)

root.mainloop()
